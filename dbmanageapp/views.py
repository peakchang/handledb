import os
import time
import urllib
from urllib.parse import urlencode, parse_qsl, urlsplit
from datetime import datetime, timedelta
from itertools import chain

import xlwt
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.utils.dateformat import DateFormat

import json
import random
import re

from django.contrib import messages
from django.db.models import Q, Sum
from django.http import JsonResponse, Http404, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect

# Create your views here.
from django.urls import reverse, reverse_lazy, resolve

from accountapp.models import User
from allmanageapp.models import AllManage
from dbmanageapp.models import MarketingList, UploadDbName, UploadDb, DbSetting, DbMemo, PaidList
from openpyxl import load_workbook

from dateutil.relativedelta import *


def div_dbname(request):
    context = {}

    # 마케팅 리스트 에러 처리 (초기값 셋팅 X)
    try:
        marketing_list = MarketingList.objects.all()
        context['marketing_list'] = marketing_list
    except:
        error = '마케팅 리스트를 추가해주세요!'
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})


    try:
        on_sd_oo = request.GET['sd']
        on_ed_oo = request.GET['ed']
        datetime_format = "%Y-%m-%d"
        on_sd = datetime.strptime(on_sd_oo, datetime_format)
        on_ed = datetime.strptime(on_ed_oo, datetime_format)
        div_date = set_search_day(on_sd, on_ed)
        context['sd'] = on_sd_oo
        context['ed'] = on_ed_oo
    except:
        now_datetime = datetime.today()
        f_datetime = datetime(now_datetime.year, now_datetime.month, 1)
        div_date = set_search_day(f_datetime, now_datetime)

    if request.method == 'POST':
        mk = request.POST.get('mk')
        context['mk'] = mk
        dbname_num = request.POST.getlist('dbname_num[]')
        dbname_id = request.POST.getlist('dbname_id[]')
        get_mk = MarketingList.objects.get(mk_company=mk)

        for val in dbname_num:
            temp_dbname = UploadDbName.objects.get(id=dbname_id[int(val)])
            temp_dbname.dbn_mkname = get_mk
            temp_dbname.save()

    db_name_list = UploadDbName.objects.filter(dbn_date__range=[div_date[0], div_date[1]])
    context['db_name_list'] = db_name_list

    return render(request, 'dbmanageapp/div_dbname.html', context)


def make_excel(request):
    sd = request.POST.get('ex_sd')
    ed = request.POST.get('ex_ed')
    ns = request.POST.get('ex_ns')
    print(ns)


    datetime_format = "%Y-%m-%d"
    onsd = datetime.strptime(sd, datetime_format)
    oned = datetime.strptime(ed, datetime_format)
    set_date = set_search_day(onsd, oned)

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'example.xls'
    wb = xlwt.Workbook(encoding='ansi')  # encoding은 ansi로 해준다.
    ws = wb.add_sheet('sheet1')  # 시트 추가

    all_memo = DbMemo.objects.filter(dm_date__range=[set_date[0], set_date[1]])

    row_num = 0
    col_names = ['마케터', '전화번호', '이름', '나이', '성별', '자산', '담당자', '담당자 닉네임', '상태', '결제금액', '결제상태']

    # 열이름을 첫번째 행에 추가 시켜준다.
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)

    # 데이터 베이스에서 유저 정보를 불러온다.

    print(set_date[0])
    print(set_date[1])
    print(ns)
    all_db_list = UploadDb.objects.select_related('db_mkname').filter(db_date__range=[set_date[0], set_date[1]], db_status=ns)
    print(all_db_list)

    rows = []
    for dblist in all_db_list:
        chk_memo = all_memo.filter(dm_chkdb=dblist)
        print(dblist)

        try:
            chk_marketer = dblist.db_mkname.mk_company
        except:
            chk_marketer = ""
        print(chk_marketer)

        print(dblist.db_phone)
        set_list = [chk_marketer, dblist.db_phone, dblist.db_member, dblist.db_age, dblist.db_sex, dblist.db_inv,
                    dblist.db_manager, dblist.db_manager_nick, dblist.db_status, dblist.db_paidprice,
                    dblist.db_paidstatus]
        if (chk_memo):
            for chkcount, memo in enumerate(chk_memo):
                if chkcount > 2:
                    break
                else:
                    set_list.append(memo.dm_memos)

        rows.append(set_list)
    # 유저정보를 한줄씩 작성한다.
    for row in rows:
        row_num += 1
        for col_num, attr in enumerate(row):
            ws.write(row_num, col_num, attr)

    wb.save(response)
    return response


def update_db(request):
    chk_db_list = DbMemo.objects.all()
    for dlist in chk_db_list:
        if dlist.dm_manager == "":
            dlist.dm_manager = dlist.dm_chkdb.db_manager
            dlist.save()
    return HttpResponse('완료가 잘 되었는지 봅시다.')


@login_required
def dbmainpage(request):
    q = Q()
    j = Q()
    if not request.user.rete == 'A':
        q.add(Q(db_manager=request.user), q.AND)
        j.add(Q(db_manager=request.user), q.AND)

    now_datetime = datetime.today()
    f_datetime = datetime(now_datetime.year, now_datetime.month, 1)

    last_month_first = datetime(now_datetime.year, now_datetime.month, 1) + relativedelta(months=-1)
    last_month_last = datetime(now_datetime.year, now_datetime.month, 1) + relativedelta(seconds=-1)

    now_month_setdate = set_search_day(f_datetime, now_datetime)
    last_month_setdate = set_search_day(last_month_first, last_month_last)

    q.add(Q(db_date__range=[now_month_setdate[0], now_month_setdate[1]]), q.AND)
    j.add(Q(db_date__range=[last_month_setdate[0], last_month_setdate[1]]), q.AND)

    now_month_price = UploadDb.objects.filter(q)
    last_month_price = UploadDb.objects.filter(j)

    n_sales = now_month_price.aggregate(Sum('db_paidprice'))
    l_sales = last_month_price.aggregate(Sum('db_paidprice'))
    n_sales = n_sales['db_paidprice__sum']
    l_sales = l_sales['db_paidprice__sum']
    i = 0

    if not n_sales or not l_sales:
        growth_per = None
    else:
        growth_per = "%.2f%%" % (n_sales / l_sales * 100.0)

    return render(request, 'dbmanageapp/mainpage.html',
                  {'n_sales': n_sales, 'l_sales': l_sales, 'growth_per': growth_per})


@login_required
def base_setting(request):
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        upload_img = request.FILES.get('logo_image')
        ds_status = request.POST.get('ds_status')
        ds_statusbase = request.POST.get('ds_statusbase')
        theme_status = request.POST.get('theme_status')
        ds_overlapallow = request.POST.get('ds_overlapallow')

        set_model = DbSetting.objects.last()

        if set_model:
            set_model.company_name = company_name
            set_model.ds_status = ds_status
            set_model.ds_statusbase = ds_statusbase
            set_model.ds_overlapallow = ds_overlapallow
            set_model.theme_status = theme_status
            if upload_img:
                if set_model.logo_image:
                    os.remove(set_model.logo_image.path)
                set_model.logo_image = upload_img
            set_model.save()
            messages.success(request, '성공적으로 반영되었습니다.')
        else:
            dbset = DbSetting()
            dbset.company_name = company_name
            dbset.ds_status = ds_status
            dbset.ds_statusbase = ds_statusbase
            dbset.ds_overlapallow = ds_overlapallow
            dbset.theme_status = theme_status
            if upload_img:
                dbset.logo_image = upload_img
            dbset.save()

    try:
        set_content = DbSetting.objects.last()
        select_menu = set_content.ds_status
        select_menu = select_menu.split(',')
        ontm = set_content.theme_status
    except:
        ontm = ""
        select_menu = ""
    return render(request, 'dbmanageapp/base_setting.html',
                  {'set_content': set_content, 'select_menu': select_menu, 'ot': ontm})


def marketer_stats(request):
    # 마케팅 리스트 에러 처리 (초기값 셋팅 X)
    try:
        marketing_list = MarketingList.objects.all()
    except:
        error = '마케팅 리스트를 추가해주세요!'
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    q = Q()
    j = Q()
    geton = get_getlist(request, q, j)

    mk_list = MarketingList.objects.all()
    mk_list_value = MarketingList.objects.values()
    ii = 0
    mk_on_list = []

    for mlist in mk_list:
        paid_a_sum = 0

        mk_on_list.append(mk_list_value[ii])
        # 검색된 날짜 기준에 등록된 DB의 전체 리스트 (전체 등록건)
        all_dblist = mlist.dbmkname.filter(db_date__range=[geton['set_date'][0], geton['set_date'][1]])

        # 검색된 날짜 기준에 등록된 DB이름에 기재된 DB 값들의 합
        seach_dbname_for_spend_db = mlist.mkdname.filter(dbn_date__range=[geton['set_date'][0], geton['set_date'][1]])
        spend_db = seach_dbname_for_spend_db.aggregate(Sum('dbn_price'))

        # 검색된 날짜 기준에 마지막 결제일이 포함되어 있는 DB의 리스트 (누적 결제건)
        paid_dblist = mlist.dbmkname.filter(db_lastpaiddate__range=[geton['set_date'][0], geton['set_date'][1]],
                                            db_paidstatus='Y')
        for dlist in paid_dblist:
            # 검색된 DB의 이번달 내지 검색된 날짜 기준으로 결제된 금액을 구함
            inner_paid_list = dlist.plchkdb.filter(pl_paiddate__range=[geton['set_date'][0], geton['set_date'][1]])
            for pilist in inner_paid_list:
                paid_a_sum = paid_a_sum + pilist.pl_paidprice

        # 누적 결제금
        all_count = all_dblist.count()
        paid_count = paid_dblist.count()
        mk_on_list[ii]['all_count'] = all_count
        mk_on_list[ii]['acc_count'] = len(paid_dblist)
        if paid_count > 0:
            mk_on_list[ii]['acc_per'] = "%.2f%%" % (paid_count / all_count * 100.0)
        else:
            mk_on_list[ii]['acc_per'] = '0%'
        mk_on_list[ii]['acc_price'] = paid_a_sum

        if paid_count > 0:
            acc_aver = paid_a_sum // paid_count
            mk_on_list[ii]['acc_aver'] = acc_aver
        else:
            mk_on_list[ii]['acc_aver'] = 0

        if not spend_db['dbn_price__sum']:
            mk_on_list[ii]['spend_db'] = 0
        else:
            mk_on_list[ii]['spend_db'] = spend_db['dbn_price__sum']

        ii += 1

    pagenum = make_get_page(mk_on_list, geton['get_page_num'], 10)
    mk_on_list = mk_on_list[pagenum[0]:pagenum[1]]

    return render(request, 'dbmanageapp/marketer_stats.html',
                  {'mk_on_list': mk_on_list, 'get_list': geton, 'pageval': pagenum[4],
                   'get_page_num': geton['get_page_num'], 'marketing_list': marketing_list})


@login_required
def alldblist(request):
    # 상태값 입력 안되었을때 에러처리
    try:
        chk_db = DbSetting.objects.last()
        all_status = chk_db.ds_status
        status_list = all_status.split(',')
        if not all_status:
            raise Http404()
    except:
        error = "상태값을 먼저 셋팅 해주세요"
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # 마케팅 리스트 에러 처리 (초기값 셋팅 X)
    try:
        marketing_list = MarketingList.objects.all()
    except:
        error = '마케팅 리스트를 추가해주세요!'
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # 담당자(매니저) 리스트 뿌려주기
    manager_list = User.objects.filter(rete='D')

    # 여기서부터 Q값 넣기 시작!!
    q = Q()
    j = Q()

    get_list = {}
    geton = get_getlist(request, q, j)
    status_count = []

    # 변경 부분1
    q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)

    all_get = UploadDb.objects.select_related('db_mkname').filter(q)
    all_count = all_get.count()

    for slist in status_list:
        status_get = UploadDb.objects.select_related('db_mkname').filter(q).filter(db_status=slist)
        status_count.append(status_get.count())

    # 전체 페이지값을 구해 페이지네이션을 구현한 뒤 원하는 갯수만큼 출력
    db_list = UploadDb.objects.select_related('db_mkname').filter(q)

    rangenum = list(reversed(range(1, db_list.count() + 1)))

    pagenum = make_get_page(db_list, geton['get_page_num'], geton['wp'])
    # 변경부분2
    db_list_val = UploadDb.objects.select_related('db_mkname').filter(q).order_by('-db_date')[pagenum[0]:pagenum[1]]
    pg_rangenum = rangenum[pagenum[0]:pagenum[1]]

    alldb_zip = zip(pg_rangenum, db_list_val)

    if request.method == 'POST':

        list_num = request.POST.getlist('listcount[]')
        list_id = request.POST.getlist('listid[]')
        change_status = request.POST.getlist('change_status[]')
        change_manager = request.POST['change_manager']
        change_manager_nick = request.POST['change_manager_nick']
        change_all_status = request.POST.get('change_all_status')

        print(change_manager)
        print(change_all_status)
        if 'update' in request.POST['submit_btn']:
            for val in list_num:
                temp_item = UploadDb.objects.get(id=list_id[int(val)])
                temp_item.db_status = change_status[int(val)]
                if change_manager:
                    temp_item.db_manager = change_manager
                    temp_item.db_manager_nick = change_manager_nick
                if change_all_status:
                    temp_item.db_status = change_all_status
                temp_item.save()
        elif 'all_delete' in request.POST['submit_btn']:
            now_datetime = datetime.today()
            set_time_today = set_search_day(now_datetime, now_datetime)
            del_alldb = UploadDb.objects.filter(db_date__range=[set_time_today[0], set_time_today[1]])
            del_alldb.delete()

        elif 'delete' in request.POST['submit_btn']:
            for val in list_num:
                temp_item = UploadDb.objects.get(id=list_id[int(val)])
                temp_item.delete()

        qstring = request.POST.get('qstring')
        response = redirect(reverse('dbmanage:alldblist'))
        response['Location'] += "?"
        response['Location'] += qstring
        return response
    return render(request, 'dbmanageapp/alldblist.html',
                  {'db_list_val': alldb_zip, 'manager_list': manager_list, 'all_status': all_status,
                   'status_list': status_list, 'status_count': status_count,
                   'marketing_list': marketing_list, 'pageval': pagenum[4],
                   'get_page_num': geton['get_page_num'], 'get_list': geton, 'all_count': all_count}, )


@login_required
def status_stats(request):
    # 상태값 입력 안되었을때 에러처리
    try:
        chk_db = DbSetting.objects.last()
        all_status = chk_db.ds_status
        status_list = all_status.split(',')
        if not all_status:
            raise Http404()
    except:
        error = "상태값을 먼저 셋팅 해주세요"
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # 마케팅 리스트 에러 처리 (초기값 셋팅 X)
    try:
        marketing_list = MarketingList.objects.all()
    except:
        error = '마케팅 리스트를 추가해주세요!'
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # 담당자(매니저) 리스트 뿌려주기
    manager_list = User.objects.filter(rete='D').values('username')

    # 여기서부터 Q값 넣기 시작!!
    q = Q()
    j = Q()

    geton = get_getlist(request, q, j)

    all_list_arr = []
    all_status_list = ['총계']
    all_status_list = all_status_list + status_list

    for manager in manager_list:
        manager_list_arr = []
        manager_list_arr.append(manager)
        q = Q()
        q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)
        q.add(Q(db_manager=manager['username']), q.AND)
        chk_db = UploadDb.objects.filter(q)
        manager_list_arr.append(chk_db.count())
        for slist in status_list:
            chk_count = chk_db.filter(db_status=slist)
            manager_list_arr.append(chk_count.count())
        all_list_arr.append(manager_list_arr)

    sum_list_arr = []
    per_list_arr = []
    for smlist in all_status_list:
        sum_list_arr.append(0)
        per_list_arr.append(0)
    #
    sum_list = ['총합']
    for ilist in all_list_arr:
        k = 0
        for bon in ilist:
            if type(bon) != int:
                continue
            else:
                sum_list_arr[k] = sum_list_arr[k] + bon
                k += 1
    sum_list = sum_list + sum_list_arr
    make_sum = sum_list[1]

    j = -1
    for perchk in sum_list_arr:
        j = j + 1
        if j == 0:
            continue
        else:
            if make_sum == 0:
                make_sum = 1
            per_on = "%.2f%%" % (perchk / make_sum * 100.0)
            per_list_arr[j] = per_on
    per_list = ['']
    per_list = per_list + per_list_arr
    appon_list = zip(per_list, sum_list)

    return render(request, 'dbmanageapp/status_stats.html',
                  {'all_list_arr': all_list_arr, 'status_list': all_status_list, 'appon_list': appon_list})


@login_required
def emp_dblist(request):
    # 여기서부터 Q값 넣기 시작!!
    q = Q()
    j = Q()
    get_list = {}
    geton = get_getlist(request, q, j)
    status_count = []
    q.add(Q(db_manager=request.user), q.AND)
    try:
        chk_db = DbSetting.objects.last()
        all_status = chk_db.ds_status
        status_list = all_status.split(',')
        if not all_status:
            raise Http404()
    except:
        error = "상태값을 먼저 셋팅 해주세요"
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # 변경 부분1
    q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)

    # q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)

    status_count = []

    all_get = UploadDb.objects.select_related('db_mkname').filter(q)
    all_count = all_get.count()
    for slist in status_list:
        status_get = UploadDb.objects.select_related('db_mkname').filter(q).filter(db_status=slist)
        status_count.append(status_get.count())

    # 전체 페이지값을 구해 페이지네이션을 구현한 뒤 원하는 갯수만큼 출력
    db_list = UploadDb.objects.select_related('db_mkname').filter(q)

    rangenum = list(reversed(range(1, db_list.count() + 1)))

    pagenum = make_get_page(db_list, geton['get_page_num'], geton['wp'])
    # 변경부분2
    db_list_val = UploadDb.objects.select_related('db_mkname').filter(q).order_by('-db_date')[pagenum[0]:pagenum[1]]
    pg_rangenum = rangenum[pagenum[0]:pagenum[1]]

    alldb_zip = zip(pg_rangenum, db_list_val)

    if request.method == 'POST':
        list_num = request.POST.getlist('listcount[]')
        list_id = request.POST.getlist('listid[]')
        change_status = request.POST.getlist('change_status[]')
        change_manager = request.POST.get('change_manager')
        change_manager_nick = request.POST.get('change_manager_nick')

        for val in list_num:
            temp_item = UploadDb.objects.get(id=list_id[int(val)])
            temp_item.db_status = change_status[int(val)]
            if change_manager:
                temp_item.db_manager = change_manager
                temp_item.db_manager_nick = change_manager_nick
            temp_item.save()

        qstring = request.POST.get('qstring')
        response = redirect(reverse('dbmanage:emp_dblist'))
        response['Location'] += "?"
        response['Location'] += qstring
        return response

    return render(request, 'dbmanageapp/emp_dblist.html',
                  {'db_list_val': alldb_zip, 'status_count': status_count, 'status_count': status_count,
                   'status_list': status_list,
                   'pageval': pagenum[4], 'get_page_num': geton['get_page_num'], 'get_list': geton,
                   'all_count': all_count}, )


@login_required
def emp_dbstats(request):
    # -------------- 뿌려주기 옵션값 끝!
    q = Q()
    j = Q()

    get_list = {}
    geton = get_getlist(request, q, j)

    q.add(Q(db_manager=request.user), q.AND)
    j.add(Q(db_manager=request.user), q.AND)

    q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)
    j.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)

    q.add(Q(db_paidstatus='Y'), q.AND)
    q.add(~Q(db_manager=''), q.AND)

    db_list = UploadDb.objects.select_related('db_mkname').filter(q)
    pagenum = make_get_page(db_list, geton['get_page_num'], geton['wp'])
    base_db = UploadDb.objects.select_related('db_mkname').filter(q).order_by('-id')[pagenum[0]:pagenum[1]]

    sum = UploadDb.objects.filter(q).aggregate(Sum('db_paidprice'))
    if not sum['db_paidprice__sum']:
        sum['db_paidprice__sum'] = 0

    return render(request, 'dbmanageapp/emp_dbstats.html',
                  {'base_db': base_db, 'pageval': pagenum[4],
                   'get_page_num': geton['get_page_num'], 'get_list': geton, 'sum': sum})


@login_required
def sale_st(request):
    ## 기본 선택값 뿌려주기
    # 담당자(매니저) 리스트 뿌려주기
    manager_list = User.objects.filter(rete='D')

    # -------------- 뿌려주기 옵션값 끝!
    q = Q()
    j = Q()

    get_list = {}
    geton = get_getlist(request, q, j)

    q.add(Q(db_lastpaiddate__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)
    j.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)

    q.add(Q(db_paidstatus='Y'), q.AND)
    q.add(~Q(db_manager=''), q.AND)

    db_list = UploadDb.objects.select_related('db_mkname').filter(q)
    pagenum = make_get_page(db_list, geton['get_page_num'], geton['wp'])
    base_db = UploadDb.objects.select_related('db_mkname').filter(q).order_by('-db_lastpaiddate')[pagenum[0]:pagenum[1]]

    sum = UploadDb.objects.filter(q).aggregate(Sum('db_paidprice'))
    if not sum['db_paidprice__sum']:
        sum['db_paidprice__sum'] = 0

    return render(request, 'dbmanageapp/sale_st.html',
                  {'base_db': base_db, 'pageval': pagenum[4],
                   'get_page_num': geton['get_page_num'], 'manager_list': manager_list,
                   'get_list': geton, 'sum': sum})


@login_required
def divdb(request):
    # 마케팅 리스트 에러 처리 (초기값 셋팅 X) + 값 뿌려주기
    try:
        marketing_list = MarketingList.objects.all()
    except:
        error = '마케팅 리스트를 추가해주세요!'
        return render(request, 'dbmanageapp/alldblist.html', {'error': error})

    # set_date = set_search_day(now_datetime, now_datetime)
    q = Q()
    # 미분배 DB 목록 수량 구하기
    db_count_arr = []

    dbn_list = UploadDbName.objects.all()

    all_db_count = 0
    for dnlist in dbn_list:
        arr = []
        db_count = UploadDb.objects.filter(db_name=dnlist, db_manager__isnull=True).count()
        if db_count > 0:
            arr.append(dnlist.dbn_mkname.mk_company)
            arr.append(dnlist.dbn_name)
            arr.append(db_count)
            arr.append(dnlist.dbn_date)
            db_count_arr.append(arr)

        all_db_count = all_db_count + db_count

    # 검색 조건들 받자
    j = Q()
    j.add(Q(db_manager__isnull=True), j.AND)
    q.add(Q(db_manager__isnull=True), j.AND)

    geton = get_getlist(request, q, j)

    q.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)
    j.add(Q(db_date__range=[geton['set_date'][0], geton['set_date'][1]]), q.AND)
    # 날짜 GET 값 받기 에러 처리
    db_list = UploadDb.objects.filter(q)
    userlist = User.objects.filter(rete='D',status='Y')


    error_text = ""

    if request.method == 'POST':

        divdb_list = request.POST.getlist('divdb[]')
        divnick_list = request.POST.getlist('divnick[]')
        divid_list = request.POST.getlist('divid[]')

        list_int = listStrToInt(divdb_list)
        sum_listint = sum(list_int)

        try:
            if sum_listint == 0 or sum_listint > db_list.count():
                raise

            k = 0
            for list in list_int:

                db_list = UploadDb.objects.filter(q)
                db_id_list = []
                for onid in db_list:
                    db_id_list.append(onid.id)

                random_val = random.sample(db_id_list, list)
                for i in random_val:
                    div_dv_update = UploadDb.objects.get(id=i)
                    div_dv_update.db_manager = divid_list[k]
                    div_dv_update.db_manager_nick = divnick_list[k]
                    div_dv_update.db_date = timezone.now()
                    div_dv_update.save()
                k += 1

            messages.success(request, "분배가 완료되었습니다.")
            return HttpResponseRedirect(reverse('dbmanage:divdb'))
        except:
            divdb_list = request.POST.getlist('divdb[]')
            list_int = listStrToInt(divdb_list)
            if sum(list_int) == 0:
                error_text = '분배할 값이 입력되지 않았습니다.'
            elif sum(list_int) > db_list.count():
                error_text = '분배할 값이 DB 수량보다 많습니다.'
    return render(request, 'dbmanageapp/divdb.html',
                  {'db_list': db_list, 'userlist': userlist, 'all_db_count': all_db_count, 'db_count_arr': db_count_arr,
                   'get_list': geton,
                   'marketing_list': marketing_list, 'error_text': error_text})


@login_required
def markerlist(request):
    if request.method == 'POST':
        btnval = request.POST['gosubmit']
        if btnval == 'create':
            newmarketer = MarketingList(mk_company=request.POST['mk_company'], mk_name=request.POST['mk_name'],
                                        mk_phone=request.POST['mk_phone'], mk_advtype=request.POST['mk_advtype'],
                                        mk_url=request.POST['mk_url'], mk_memo=request.POST['mk_memo'])
            newmarketer.save()
        elif btnval == 'update':
            mk_id = request.POST['mk_id']
            updatemarketer = MarketingList.objects.get(id=mk_id)
            updatemarketer.mk_company = request.POST['mk_company']
            updatemarketer.mk_name = request.POST['mk_name']
            updatemarketer.mk_phone = request.POST['mk_phone']
            updatemarketer.mk_advtype = request.POST['mk_advtype']
            updatemarketer.mk_url = request.POST['mk_url']
            updatemarketer.mk_status = request.POST['mk_status']
            updatemarketer.save()
        else:
            mk_id = request.POST['mk_id']
            updatemarketer = MarketingList.objects.get(id=mk_id)
            updatemarketer.delete()

    marketing_list = MarketingList.objects.all()
    return render(request, 'dbmanageapp/marketinglist.html', {'marketing_list': marketing_list})


@login_required
def newdbup(request):
    try:
        marketing_list = MarketingList.objects.all()
    except:
        marketing_list = '마케팅 리스트를 추가해주세요!'

    try:
        sample_list = AllManage.objects.last()
    except:
        sample_list = ""

    overlap_db = []

    if request.method == 'POST':

        print("asldijfalisjdfliajsdlfjaslidfjalisdjfasfd")
        now = datetime.now()
        before_three_week = now - timedelta(weeks=10)
        set_tr_date = set_search_day(before_three_week, now)
        overlap_count = 0

        dblist_text = request.POST['dblist_text']
        new_db_file = request.FILES.get('dblist_file')
        back_db_file = request.FILES.get('backup_dblist_file')

        print(dblist_text)
        print(new_db_file)
        print(back_db_file)
        if dblist_text or new_db_file:
            if dblist_text and new_db_file is None:
                dblist_text = dblist_text.splitlines(False)

                i = 0
                dblist = []
                for val in dblist_text:
                    val = re.sub("\!|\'|\?|\-", "", val)
                    val = val.split(',')
                    if val[0].isdigit():
                        val[0] = val[0].zfill(11)
                    if len(val) == 1:
                        val.append(val[0])
                    dblist.append(val)
            else:
                if new_db_file is not None:
                    files = request.FILES.get('dblist_file')
                    try:
                        load_wb = load_workbook(files, data_only=True)
                    except:
                        error_message = "엑셀파일에 문제가 있습니다. 새로운 엑셀파일에 데이터를 넣고 업로드 해주세요"
                        return render(request, 'dbmanageapp/newdbup.html',
                                      {'marketing_list': marketing_list, 'sample_list': sample_list,
                                       'error_message': error_message})
                    load_ws = load_wb['Sheet1']
                    dblist = []
                    for row in load_ws.rows:
                        row_value = []

                        for cell in row:
                            cellval = cell.value

                            cellval = str(cellval)
                            cellval = re.sub("\!|\'|\?|\-", "", cellval)

                            if cellval.isdigit():
                                cellval = cellval.zfill(11)
                            if cellval == 'None' or not cellval:
                                cellval = ""
                            row_value.append(cellval)

                        chk_list = [v for v in row_value if v]

                        if not chk_list:
                            break

                        # if not row_value[1]:
                        #     row_value[1] = row_value[0]
                        if len(row_value) < 2:
                            temp_rowval = row_value[0]
                            row_value.append(temp_rowval)
                        elif not row_value[1]:
                            row_value[1] = row_value[0]

                        dblist.append(row_value)

            try:
                base_seton = DbSetting.objects.last()
                base_set_list = base_seton.ds_status.split(',')
                base_status = base_set_list[0]
                overlap_allow = base_seton.ds_overlapallow.split(',')

                # 업로드 DB 구분을 위한 이름을 만듦
                dbn_mkname = request.POST.get('dbn_mkname')
                dbn_name = request.POST.get('dbn_name')
                dbn_price = request.POST.get('dbn_price')
                dbn_price = re.sub("\!|\'|\,|\-", "", dbn_price)
                dbn_price = int(dbn_price)
                dbn_memo = request.POST.get('dbn_memo')
                chk_ml = MarketingList.objects.get(mk_company=dbn_mkname)
                dbn_items = UploadDbName(dbn_mkname=chk_ml, dbn_name=dbn_name, dbn_price=dbn_price, dbn_memo=dbn_memo)
                dbn_items.save()

                temp_udb = UploadDbName.objects.last()
                onfr = UploadDbName.objects.get(id=temp_udb.id)
                temp_mkt = MarketingList.objects.get(mk_company=dbn_mkname)

                # DB에 값 넣기

                # 중복항목 검사할 전체 쿼리 값
                chk_db_list = UploadDb.objects.filter(db_date__range=[set_tr_date[0], set_tr_date[1]])
                overlap_count = 0

                for dbval in dblist:
                    if len(dbval) < 6:
                        set_arr_count = 6 - len(dbval)
                        for i in range(set_arr_count):
                            dbval.append('')
                    # try:
                    #     if chk_db_list.get(db_phone=dbval[0]):
                    #         overlap_count += 1
                    #         continue
                    # except:
                    #     pass
                    try:
                        chk_db = chk_db_list.filter(db_phone=dbval[0]).last()
                        if chk_db:
                            if chk_db.db_status in overlap_allow:
                                chk_db_on = chk_db_list.filter(db_phone=dbval[0])
                                for del_db in chk_db_on:
                                    del_db.delete()
                                pass
                            else:
                                overlap_count += 1
                                continue
                    except:
                        pass

                    db_up = UploadDb(db_name=onfr, db_mkname=temp_mkt, db_member=dbval[1], db_phone=dbval[0],
                                     db_age=dbval[2], db_sex=dbval[3], db_inv=dbval[4], db_status=base_status)
                    db_up.save()

                    if dbval[5]:
                        serch_menodb = UploadDb.objects.last()
                        memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[5])
                        memoup.save()

                if len(dblist) == overlap_count:
                    temp_udb.delete()
                    overlap = "모든 항목이 중복됩니다. 같은 파일이 업로드 된것 같습니다."
                elif overlap_count:
                    overlap = f"{overlap_count} 건이 중복되었습니다."
                else:
                    overlap = ""
                messages.success(request, f"DB 업로드가 완료 되었습니다. {overlap}")
            except:
                error_message = "업로드 요청된 DB가 없습니다. DB를 입력해주세요"
                return render(request, 'dbmanageapp/newdbup.html',
                              {'marketing_list': marketing_list, 'sample_list': sample_list,
                               'error_message': error_message})
        elif back_db_file:
            try:
                load_wb = load_workbook(back_db_file, data_only=True)
            except:
                error_message = "엑셀파일에 문제가 있습니다. 새로운 엑셀파일에 데이터를 넣고 업로드 해주세요"
                return render(request, 'dbmanageapp/newdbup.html',
                              {'marketing_list': marketing_list, 'sample_list': sample_list,
                               'error_message': error_message})
            load_ws = load_wb['Sheet1']
            dblist = []
            for row in load_ws.rows:
                row_value = []

                for cell in row:
                    cellval = cell.value
                    cellval = str(cellval)
                    cellval = re.sub("\!|\'|\?|\-", "", cellval)

                    if cellval.isdigit():
                        cellval = cellval.zfill(11)
                    if cellval == 'None' or not cellval:
                        cellval = ""
                    row_value.append(cellval)

                chk_list = [v for v in row_value if v]

                if not chk_list:
                    break
                dblist.append(row_value)

            chk_db_list = UploadDb.objects.filter(db_date__range=[set_tr_date[0], set_tr_date[1]])
            overlap_count = 0
            for dbval in dblist:
                chk_db = chk_db_list.filter(db_phone=dbval[0]).last()
                if chk_db:
                    pass

                db_up = UploadDb(db_phone=dbval[1], db_member=dbval[2], db_age=dbval[3], db_sex=dbval[4], db_inv=dbval[5], db_manager=dbval[6], db_manager_nick=dbval[7], db_status=dbval[8], db_paidprice=dbval[9], db_paidstatus=dbval[10])
                db_up.save()

                if dbval[11] != "":
                    serch_menodb = UploadDb.objects.last()
                    memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[11])
                    memoup.save()

                try:
                    if dbval[12] != "":
                        serch_menodb = UploadDb.objects.last()
                        memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[12])
                        memoup.save()
                except:
                    pass

                try:
                    if dbval[13] != "":
                        serch_menodb = UploadDb.objects.last()
                        memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[13])
                        memoup.save()
                except:
                    pass

    return render(request, 'dbmanageapp/newdbup.html', {'marketing_list': marketing_list, 'sample_list': sample_list})


# def new_dbup(request):
#     try:
#         marketing_list = MarketingList.objects.all()
#     except:
#         marketing_list = '마케팅 리스트를 추가해주세요!'
#
#     try:
#         sample_list = AllManage.objects.last()
#     except:
#         sample_list = ""
#
#     overlap_db = []
#
#     if request.method == 'POST':
#         now = datetime.now()
#         before_three_week = now - timedelta(weeks=10)
#         set_tr_date = set_search_day(before_three_week, now)
#         overlap_count = 0
#
#         dblist_text = request.POST['dblist_text']
#         new_db_file = request.FILES.get('dblist_file')
#         back_db_file = request.FILES.get('backup_dblist_file')
#
#         if dblist_text and new_db_file is None:
#             dblist_text = dblist_text.splitlines(False)
#
#             i = 0
#             dblist = []
#             for val in dblist_text:
#                 val = re.sub("\!|\'|\?|\-", "", val)
#                 val = val.split(',')
#                 if val[0].isdigit():
#                     val[0] = val[0].zfill(11)
#                 if len(val) == 1:
#                     val.append(val[0])
#                 dblist.append(val)
#         else:
#             if new_db_file is not None:
#                 files = request.FILES.get('dblist_file')
#                 try:
#                     load_wb = load_workbook(files, data_only=True)
#                 except:
#                     error_message = "엑셀파일에 문제가 있습니다. 새로운 엑셀파일에 데이터를 넣고 업로드 해주세요"
#                     return render(request, 'dbmanageapp/newdbup.html',
#                                   {'marketing_list': marketing_list, 'sample_list': sample_list,
#                                    'error_message': error_message})
#                 load_ws = load_wb['Sheet1']
#                 dblist = []
#                 for row in load_ws.rows:
#                     row_value = []
#
#                     for cell in row:
#                         cellval = cell.value
#
#                         cellval = str(cellval)
#                         cellval = re.sub("\!|\'|\?|\-", "", cellval)
#
#                         if cellval.isdigit():
#                             cellval = cellval.zfill(11)
#                         if cellval == 'None' or not cellval:
#                             cellval = ""
#                         row_value.append(cellval)
#
#                     chk_list = [v for v in row_value if v]
#
#                     if not chk_list:
#                         break
#
#                     if len(row_value) < 2:
#                         temp_rowval = row_value[0]
#                         row_value.append(temp_rowval)
#                     elif not row_value[1]:
#                         row_value[1] = row_value[0]
#
#                     dblist.append(row_value)
#
#         try:
#             base_seton = DbSetting.objects.last()
#             base_set_list = base_seton.ds_status.split(',')
#             base_status = base_set_list[0]
#             overlap_allow = base_seton.ds_overlapallow.split(',')
#
#             # 업로드 DB 구분을 위한 이름을 만듦
#             dbn_mkname = request.POST.get('dbn_mkname')
#             dbn_name = request.POST.get('dbn_name')
#             dbn_price = request.POST.get('dbn_price')
#             dbn_price = re.sub("\!|\'|\,|\-", "", dbn_price)
#             dbn_price = int(dbn_price)
#             dbn_memo = request.POST.get('dbn_memo')
#             chk_ml = MarketingList.objects.get(mk_company=dbn_mkname)
#             dbn_items = UploadDbName(dbn_mkname=chk_ml, dbn_name=dbn_name, dbn_price=dbn_price, dbn_memo=dbn_memo)
#             dbn_items.save()
#
#             temp_udb = UploadDbName.objects.last()
#             onfr = UploadDbName.objects.get(id=temp_udb.id)
#             temp_mkt = MarketingList.objects.get(mk_company=dbn_mkname)
#
#             # DB에 값 넣기
#
#             # 중복항목 검사할 전체 쿼리 값
#             chk_db_list = UploadDb.objects.filter(db_date__range=[set_tr_date[0], set_tr_date[1]])
#             overlap_count = 0
#
#             for dbval in dblist:
#                 if len(dbval) < 6:
#                     set_arr_count = 6 - len(dbval)
#                     for i in range(set_arr_count):
#                         dbval.append('')
#
#                 try:
#                     chk_db = chk_db_list.filter(db_phone=dbval[0]).last()
#                     if chk_db:
#                         if chk_db.db_status in overlap_allow:
#                             chk_db_on = chk_db_list.filter(db_phone=dbval[0])
#                             for del_db in chk_db_on:
#                                 del_db.delete()
#                             pass
#                         else:
#                             overlap_count += 1
#                             continue
#                 except:
#                     pass
#
#                 db_up = UploadDb(db_name=onfr, db_mkname=temp_mkt, db_member=dbval[1], db_phone=dbval[0],
#                                  db_age=dbval[2], db_sex=dbval[3], db_inv=dbval[4], db_status=base_status)
#                 db_up.save()
#
#                 if dbval[5]:
#                     serch_menodb = UploadDb.objects.last()
#                     memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[5])
#                     memoup.save()
#
#             if len(dblist) == overlap_count:
#                 temp_udb.delete()
#                 overlap = "모든 항목이 중복됩니다. 같은 파일이 업로드 된것 같습니다."
#             elif overlap_count:
#                 overlap = f"{overlap_count} 건이 중복되었습니다."
#             else:
#                 overlap = ""
#             messages.success(request, f"DB 업로드가 완료 되었습니다. {overlap}")
#         except:
#             error_message = "업로드 요청된 DB가 없습니다. DB를 입력해주세요"
#             return render(request, 'dbmanageapp/newdbup.html',
#                           {'marketing_list': marketing_list, 'sample_list': sample_list,
#                            'error_message': error_message})
#
#     return render(request, 'dbmanageapp/newdbup.html', {'marketing_list': marketing_list, 'sample_list': sample_list})


def new_dbup(request):
    try:
        marketing_list = MarketingList.objects.all()
    except:
        marketing_list = '마케팅 리스트를 추가해주세요!'

    try:
        sample_list = AllManage.objects.last()
    except:
        sample_list = ""

    overlap_db = []

    if request.method == 'POST':
        now = datetime.now()
        before_three_week = now - timedelta(weeks=10)
        set_tr_date = set_search_day(before_three_week, now)
        overlap_count = 0

        dblist_text = request.POST['dblist_text']
        new_db_file = request.FILES.get('dblist_file')
        back_db_file = request.FILES.get('backup_dblist_file')

        if dblist_text and new_db_file is None:
            dblist_text = dblist_text.splitlines(False)

            i = 0
            dblist = []
            for val in dblist_text:
                val = re.sub("\!|\'|\?|\-", "", val)
                val = val.split(',')
                if val[0].isdigit():
                    val[0] = val[0].zfill(11)
                if len(val) == 1:
                    val.append(val[0])
                dblist.append(val)
        else:
            if new_db_file is not None:
                files = request.FILES.get('dblist_file')
                try:
                    load_wb = load_workbook(files, data_only=True)
                except:
                    error_message = "엑셀파일에 문제가 있습니다. 새로운 엑셀파일에 데이터를 넣고 업로드 해주세요"
                    return render(request, 'dbmanageapp/newdbup.html',
                                  {'marketing_list': marketing_list, 'sample_list': sample_list,
                                   'error_message': error_message})
                load_ws = load_wb['Sheet1']
                dblist = []
                for row in load_ws.rows:
                    row_value = []

                    for cell in row:
                        cellval = cell.value

                        cellval = str(cellval)
                        cellval = re.sub("\!|\'|\?|\-", "", cellval)

                        if cellval.isdigit():
                            cellval = cellval.zfill(11)
                        if cellval == 'None' or not cellval:
                            cellval = ""
                        row_value.append(cellval)

                    chk_list = [v for v in row_value if v]

                    if not chk_list:
                        break

                    if len(row_value) < 2:
                        temp_rowval = row_value[0]
                        row_value.append(temp_rowval)
                    elif not row_value[1]:
                        row_value[1] = row_value[0]

                    dblist.append(row_value)
        try:
            base_seton = DbSetting.objects.last()
            base_set_list = base_seton.ds_status.split(',')
            base_status = base_set_list[0]
            overlap_allow = base_seton.ds_overlapallow.split(',')

            # 업로드 DB 구분을 위한 이름을 만듦
            dbn_mkname = request.POST.get('dbn_mkname')
            dbn_name = request.POST.get('dbn_name')
            dbn_price = request.POST.get('dbn_price')
            dbn_price = re.sub("\!|\'|\,|\-", "", dbn_price)
            dbn_price = int(dbn_price)
            dbn_memo = request.POST.get('dbn_memo')
            chk_ml = MarketingList.objects.get(mk_company=dbn_mkname)
            dbn_items = UploadDbName(dbn_mkname=chk_ml, dbn_name=dbn_name, dbn_price=dbn_price, dbn_memo=dbn_memo)
            dbn_items.save()

            temp_udb = UploadDbName.objects.last()
            onfr = UploadDbName.objects.get(id=temp_udb.id)
            temp_mkt = MarketingList.objects.get(mk_company=dbn_mkname)

            # DB에 값 넣기

            # 중복항목 검사할 전체 쿼리 값
            chk_db_list = UploadDb.objects.filter(db_date__range=[set_tr_date[0], set_tr_date[1]])
            overlap_count = 0

            for dbval in dblist:
                if len(dbval) < 6:
                    set_arr_count = 6 - len(dbval)
                    for i in range(set_arr_count):
                        dbval.append('')

                try:
                    chk_db = chk_db_list.filter(db_phone=dbval[0]).last()
                    if chk_db:
                        if chk_db.db_status in overlap_allow:
                            chk_db_on = chk_db_list.filter(db_phone=dbval[0])
                            for del_db in chk_db_on:
                                del_db.delete()
                            pass
                        else:
                            overlap_count += 1
                            continue
                except:
                    pass

                db_up = UploadDb(db_name=onfr, db_mkname=temp_mkt, db_member=dbval[1], db_phone=dbval[0],
                                 db_age=dbval[2], db_sex=dbval[3], db_inv=dbval[4], db_status=base_status)
                db_up.save()

                if dbval[5]:
                    serch_menodb = UploadDb.objects.last()
                    memoup = DbMemo(dm_chkdb=serch_menodb, dm_memos=dbval[5])
                    memoup.save()

            if len(dblist) == overlap_count:
                temp_udb.delete()
                overlap = "모든 항목이 중복됩니다. 같은 파일이 업로드 된것 같습니다."
            elif overlap_count:
                overlap = f"{overlap_count} 건이 중복되었습니다."
            else:
                overlap = ""
            messages.success(request, f"DB 업로드가 완료 되었습니다. {overlap}")
        except:
            error_message = "업로드 요청된 DB가 없습니다. DB를 입력해주세요"
            return render(request, 'dbmanageapp/newdbup.html',
                          {'marketing_list': marketing_list, 'sample_list': sample_list,
                           'error_message': error_message})

    return render(request, 'dbmanageapp/newdbup.html', {'marketing_list': marketing_list, 'sample_list': sample_list})
# **********************************


def accountmanagement(request):
    if request.method == 'POST':
        id_list = request.POST.getlist('idlist[]')
        id_count = request.POST.getlist('idcount[]')
        manager_status = request.POST.getlist('manager_status[]')
        manager_rate = request.POST.getlist('manager_rate[]')
        manager_nick = request.POST.getlist('manager_nick[]')

        i = 0
        for val in id_count:
            temp_user = User.objects.get(id=id_list[int(val)])
            temp_user.rete = manager_rate[int(val)]
            temp_user.status = manager_status[int(val)]
            temp_user.nickname = manager_nick[int(val)]
            temp_user.save()

            i += 1

    user_list = User.objects.all().order_by('-id')

    return render(request, 'dbmanageapp/accountmanagement.html', {'user_list': user_list})


def detail_customer(request, id):
    try:
        # 상태값 뿌려주기
        chk_db = DbSetting.objects.values_list()

        status_list = chk_db[0][1].split(',')
        # 회원정보 뿌려주기
        customer_info = UploadDb.objects.get(id=id)
        # 댓글 리스트 뿌려주기

        memo_list = DbMemo.objects.select_related('dm_chkdb').filter(dm_chkdb=customer_info).order_by('-id')
    except:
        pass
    db_status = UploadDb.objects.get(id=id)

    if request.method == 'POST':
        if request.POST['sbm_button'] == 'update':

            status_sel = request.POST.get('status_sel')
            payment_sel = request.POST.get('paystatus_sel')
            customer_name = request.POST.get('customer_name')
            db_memo = request.POST.get('db_memo')
            db_manager = request.POST.get('ondb_manager')

            if db_memo:
                DbMemo.objects.create(dm_chkdb=db_status, dm_memos=db_memo, dm_manager=db_manager)

            db_status.db_status = status_sel
            db_status.db_paidstatus = payment_sel
            db_status.db_member = customer_name

            db_status.save()

            messages.success(request, '정상적으로 업데이트 되었습니다.')
            return HttpResponseRedirect(reverse('dbmanage:detail_customer', kwargs={'id': id}))
        elif request.POST['sbm_button'] == 'payment':

            payment_val = request.POST['db_payment']
            payment_val = re.sub(",", "", payment_val)
            payment_val = int(payment_val)

            paidlist = PaidList()
            paidlist.pl_paidprice = payment_val
            paidlist.pl_chkdb = db_status
            paidlist.save()

            db_status.db_paidstatus = 'Y'
            set_paidstatus = DbSetting.objects.last()

            db_status.db_status = set_paidstatus.ds_statusbase
            sum = db_status.plchkdb.aggregate(Sum('pl_paidprice'))
            db_status.db_paidprice = sum['pl_paidprice__sum']
            temp_now_date = timezone.now()
            db_status.db_lastpaiddate = temp_now_date
            db_status.save()
            messages.success(request, '결제값이 업데이트 되었습니다.')
            return HttpResponseRedirect(reverse('dbmanage:detail_customer', kwargs={'id': id}))

    basedb_paidlist = db_status.plchkdb.all()

    sum = db_status.plchkdb.aggregate(Sum('pl_paidprice'))

    return render(request, 'dbmanageapp/detail_customer.html',
                  {'customer_info': customer_info, 'status_list': status_list, 'memo_list': memo_list,
                   'basedb_paidlist': basedb_paidlist, 'sum': sum})


def workAjax(request):
    jsonObject = json.loads(request.body)
    if 'status_list' in jsonObject:
        bring_status = jsonObject.get('status_list')
        set_db = DbSetting.objects.all()
        if set_db.count() == 0:
            setting_status = DbSetting(ds_status=bring_status)
            setting_status.save()
        else:
            chk_db = DbSetting.objects.values_list()
            setdb_key = chk_db[0][0]
            update_set_status = DbSetting.objects.get(id=setdb_key)
            update_set_status.ds_status = bring_status
            update_set_status.save()

    elif 'db_memo' in jsonObject:
        bring_memo = jsonObject.get('db_memo')
        bring_id = jsonObject.get('now_id')
        chk_uploadub = UploadDb.objects.get(id=bring_id)
        set_dbmemo = DbMemo(dm_chkdb=chk_uploadub, dm_memos=bring_memo)
        set_dbmemo.save()

    elif 'mk_name' in jsonObject:
        bring_mkname = jsonObject.get('mk_name')
        bring_sd = jsonObject.get('date_sd')
        bring_sd = datetime.strptime(bring_sd, '%Y-%m-%d')
        bring_ed = jsonObject.get('date_ed')
        bring_ed = datetime.strptime(bring_ed, '%Y-%m-%d')
        set_date = set_search_day(bring_sd, bring_ed)

        mk_name = MarketingList.objects.get(mk_company=bring_mkname)

        chk_dbname = UploadDbName.objects.filter(dbn_mkname=mk_name, dbn_date__range=[set_date[0], set_date[1]])
        dbname_arr = []
        for dbname in chk_dbname:
            dbname_arr.append(dbname.dbn_name)

        jsonObject = {'db_names': dbname_arr}
    elif 'del_val' in jsonObject:
        detail_id = jsonObject.get('detail_id')

        temp_paid = PaidList.objects.get(id=detail_id)

        may_del_paidprice = int(temp_paid.pl_paidprice)
        may_change_db_id = temp_paid.pl_chkdb.id
        changes_db = UploadDb.objects.get(id=may_change_db_id)
        temp_price = int(changes_db.db_paidprice)
        temp_price = temp_price - may_del_paidprice
        changes_db.db_paidprice = temp_price

        if temp_price < 1:
            changes_db.db_paidstatus = 'N'
            changes_db.db_status = ""

        changes_db.save()

        temp_paid.delete()



    elif 'change_price' in jsonObject:
        detail_id = jsonObject.get('detail_id')
        change_price = jsonObject.get('change_price')
        update_item = PaidList.objects.get(id=detail_id)
        update_item.pl_paidprice = change_price
        update_item.save()

    elif 'add_username' in jsonObject:
        add_username = jsonObject.get('add_username')
        add_nickname = jsonObject.get('add_nickname')
        add_password = jsonObject.get('add_password')

        try:
            User.objects.create_user(username=add_username, password=add_password, nickname=add_nickname, )
            context = {'success_msg': '정상적으로 가입 되었습니다.'}
            return JsonResponse(context)
        except:
            return {'error!!'}
    elif 'change_pw_input' in jsonObject:
        context = {'chdii': '비밀번호 변경!!'}

        receive_num = jsonObject.get('choices_num')
        receive_pwd = jsonObject.get('change_pw_input')
        user = User.objects.get(id=int(receive_num))
        user.set_password(receive_pwd)
        user.save()
        return JsonResponse(context)
    elif 'del_memo_id' in jsonObject:
        receive_del_memo_id = jsonObject.get('del_memo_id')
        memos = DbMemo.objects.get(id=receive_del_memo_id)
        memos.delete()
    else:
        # Http404()
        context = {'testpppp': 'testkjkasjdfkajsdkfj'}
        return JsonResponse(context)

    return JsonResponse(jsonObject)


def test_chk(requese):
    # testitem = UploadDb.objects.select_related('db_name').get(id=869)
    # testitem = UploadDb.objects.get(id=869)
    # chk_mk = MarketingList.objects.get(id=6)

    test1 = MarketingList.objects.get(id=5)
    test1['testval'] = 'asdjflkasjdjsdf'
    # test2 = test1.mkdname.get(id=35)
    # test3 = test2.dbname.all()
    #
    # test4 = MarketingList.objects.all()
    # for list in test4:
    #     for inlist in list.mkdname.all():
    #         inlist.dbname.all()

    return render(requese, 'dbmanageapp/test_chk.html', )


# ************** custom 함수

def get_getlist(request, q, j):
    get_list = {}

    # 현재 상태 받기 에러처리!!
    try:
        ps = request.GET['ps']
        get_list['ps'] = ps
        if ps:
            q.add(Q(db_phone__icontains=ps), q.AND)

    except:
        get_list['ps'] = ''

    # DB 이름 받기 에러처리!!
    try:
        dn = request.GET['dn']
        get_list['dn'] = dn
        temp_dbnname = UploadDbName.objects.get(dbn_name=dn)
        q.add(Q(db_name=temp_dbnname), q.AND)
        j.add(Q(db_name=temp_dbnname), q.AND)
    except:
        get_list['dn'] = ''

    # 담당자 (매니저) 받기 에러처리!
    try:
        ns = request.GET['ns']
        get_list['ns'] = ns
        if ns:
            q.add(Q(db_status=ns), q.AND)
    except:
        get_list['ns'] = ''

    # 담당자 (매니저) 받기 에러처리!
    try:
        mn = request.GET['mn']
        get_list['mn'] = mn
        if mn:
            q.add(Q(db_manager=mn), q.AND)
    except:
        get_list['mn'] = ''

    # 페이지 개수 받기 에러처리
    try:
        wp = request.GET['wp']
        wp = int(wp)
    except:
        wp = 20
    get_list['wp'] = wp

    # 마케터 받기 에러처리
    try:
        mk = request.GET['mk']
        get_list['mk'] = mk
        chk_ml = MarketingList.objects.get(mk_company=mk)
        q.add(Q(db_mkname=chk_ml), q.AND)
    except:
        get_list['mk'] = ''

    # 현재 페이지 값 받기 에러처리
    try:
        get_page_num = request.GET['page']
        get_page_num = int(get_page_num)
    except:
        get_page_num = 1

    get_list['get_page_num'] = get_page_num

    # 날짜 받기 에러처리!
    try:
        sd = request.GET['sd']
        get_list['sd'] = sd
        sd = datetime.strptime(sd, '%Y-%m-%d')
        ed = request.GET['ed']
        get_list['ed'] = ed
        ed = datetime.strptime(ed, '%Y-%m-%d')
        set_date = set_search_day(sd, ed)
    except:
        now_datetime = datetime.today()
        # 이번달 초로 날짜 구함
        # f_datetime = datetime(now_datetime.year, now_datetime.month, 1)

        set_date = set_search_day(now_datetime, now_datetime)
    get_list['set_date'] = set_date

    get_list['q'] = q
    return get_list


def make_get_page(alllist, get_page_num, wantpage):
    # param1 : 전체 게시물 갯수 / param2 : 현재 페이지
    # 전체 카운트를 구해 원하는 숫자로 쪼개 페이지를 나눈다.

    pagecount = divmod(len(alllist), wantpage)
    temp_count = 0
    pagelist = []
    while temp_count <= pagecount[0]:
        temp_count = temp_count + 1
        pagelist.append(temp_count)

    get_page_num = int(get_page_num)
    startpage = (get_page_num - 1) * wantpage
    searchpage = get_page_num * wantpage

    get_last_page_num = int(pagecount[0])
    get_last_page_num = get_last_page_num + 1

    start_page_num = int(get_page_num) - 5
    end_page_num = int(get_page_num) + 5

    if start_page_num <= 0:
        pageval = pagelist[0:10]
    elif end_page_num >= get_last_page_num:
        fullpageval = get_last_page_num
        startval = fullpageval - 10
        pageval = pagelist[startval:fullpageval]
    else:
        pageval = pagelist[start_page_num:end_page_num]
    # 페이지 시작 값 / 몇개 노출할건지? / 전체 페이지 리스트 / 뭐더라?? 안씀 / 필터링 된 페이지 리스트
    return [startpage, searchpage, pagelist, pagecount, pageval]


def set_search_day(start_time, end_time):
    start_of_datetime = start_time.replace(hour=0, minute=0, second=0, microsecond=000000)

    end_of_day_datetime = end_time.replace(hour=23, minute=59, second=59, microsecond=999999)
    result = [start_of_datetime, end_of_day_datetime]
    return result


def listStrToInt(list_str):
    i = 0
    for list in list_str:
        if list == '':
            del list_str[i]
            list_str.insert(i, '0')
        i += 1
    list_int = [int(i) for i in list_str]
    return list_int
